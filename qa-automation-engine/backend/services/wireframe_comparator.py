"""Wireframe-to-O9-Report comparison engine.

Parses both a wireframe XLSX and an O9 report JSON into a common normalized
format, then performs a bidirectional field-level comparison to flag all
inconsistencies — regardless of which side is wrong.
"""

import io
import math
import re
from dataclasses import dataclass, field, asdict
from typing import Any, Optional

import openpyxl


# ---------------------------------------------------------------------------
# Normalized data models
# ---------------------------------------------------------------------------

@dataclass
class FilterItem:
    attribute: str
    dimension: str
    filter_type: str = ""
    single_select: Optional[bool] = None
    default_value: Optional[str] = None
    hide_dummy_members: Optional[bool] = None
    calendar_selection: Optional[bool] = None
    sort_by: Optional[str] = None
    sort_order: Optional[str] = None


@dataclass
class DimensionItem:
    attribute: str
    dimension: str
    axis: str = ""
    visible: Optional[bool] = None
    required: Optional[bool] = None
    show_subtotal: Optional[bool] = None
    show_conditional_format: Optional[bool] = None
    display_name: Optional[str] = None
    sort_by: Optional[str] = None
    item_type: str = ""  # "Level Attribute", etc.


@dataclass
class MeasureItem:
    name: str
    visible: Optional[bool] = None
    editable: Optional[bool] = None
    render_type: Optional[str] = None
    alignment: Optional[str] = None
    format_string: Optional[str] = None
    time_horizon: Optional[str] = None
    sort: Optional[str] = None
    show_local_timezone: Optional[bool] = None


@dataclass
class NormalizedReport:
    source: str  # "wireframe" or "o9_report"
    report_name: str = ""
    filters: list = field(default_factory=list)
    dimensions: list = field(default_factory=list)
    measures: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Value helpers
# ---------------------------------------------------------------------------

def _to_bool(val: Any) -> Optional[bool]:
    """Normalize a value to bool. Returns None if indeterminate."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        s = val.strip().upper()
        if s in ("TRUE", "YES", "1"):
            return True
        if s in ("FALSE", "NO", "0"):
            return False
        if s == "":
            return None
    return None


def _to_str(val: Any) -> Optional[str]:
    """Normalize a value to string. None/empty → None."""
    if val is None:
        return None
    if isinstance(val, bool):
        return str(val)
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return None
        if val == int(val):
            return str(int(val))
        return str(val)
    s = str(val).strip()
    return s if s else None


def _norm_align(val: Optional[str]) -> Optional[str]:
    """Normalize alignment strings: 'Middle Right' → 'middleright'."""
    if val is None:
        return None
    return re.sub(r"\s+", "", val).lower() or None


def _norm_axis(val: Optional[str]) -> Optional[str]:
    """Normalize axis: 'Row' → 'row', 'Column' → 'column'."""
    if val is None:
        return None
    return val.strip().lower() or None


# ---------------------------------------------------------------------------
# Wireframe XLSX parser (purpose-built, NOT using generic wireframe_parser)
# ---------------------------------------------------------------------------

_REAL_SECTION_NAMES = {
    "filter aop", "filter_aop", "filters",
    "dimension", "dimensions",
    "measure", "measures",
}


def _cell_val(ws, row, col):
    """Get cell value, stripping strings."""
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _has_fill(cell) -> bool:
    """Check if cell has non-default background fill."""
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return False
    c = fill.fgColor
    if c.type == "theme" and c.theme is not None:
        return True
    if c.type == "rgb" and c.rgb and c.rgb not in ("00000000", "FFFFFFFF"):
        return True
    if c.type == "indexed" and c.indexed is not None and c.indexed not in (0, 64):
        return True
    return False


def _is_wide_merge(ws, row_num, min_span=4) -> bool:
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row_num <= mr.max_row:
            if (mr.max_col - mr.min_col + 1) >= min_span:
                return True
    return False


def _detect_real_banners(ws, max_row, max_col):
    """Find REAL section banners — colored/merged rows with known section names."""
    banners = []
    for r in range(1, max_row + 1):
        # Count non-empty cells
        non_empty = 0
        first_val = None
        first_cell = None
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is not None and (not isinstance(v, str) or v.strip()):
                non_empty += 1
                if first_val is None:
                    first_val = str(v).strip()
                    first_cell = cell

        if non_empty == 0 or non_empty > 2 or first_val is None:
            continue

        has_color = _has_fill(first_cell) if first_cell else False
        has_merge = _is_wide_merge(ws, r, min_span=4)

        if has_color or has_merge:
            # Check if it's a known real section
            name_lower = first_val.lower().strip()
            if any(name_lower.startswith(s) or s.startswith(name_lower)
                   for s in _REAL_SECTION_NAMES):
                banners.append((r, first_val))

    return banners


def _find_header_row(ws, start, end, max_col):
    """Find header row = row with most text cells."""
    best_row, best_count = None, 0
    for r in range(start, end + 1):
        count = sum(
            1 for c in range(1, max_col + 1)
            if isinstance(ws.cell(row=r, column=c).value, str)
            and ws.cell(row=r, column=c).value.strip()
        )
        if count > best_count:
            best_count = count
            best_row = r
    return best_row if best_count >= 2 else None


def _read_headers(ws, row, max_col):
    """Read header names and their column indices."""
    headers = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = c
    return headers


def _parse_filters_from_sheet(ws, start_row, end_row, max_col, headers):
    """Parse the filter area with sub-type label awareness.

    Key behaviors:
    - Rows with a Filter Type label but no Attribute → sub-type label (skip)
    - Rows with an Attribute value → actual filter data row
    - Dimension is NOT inherited across rows; each row must have its own
    - Default Value falls back to Member Filter column (wireframe quirk)
    """
    filters = []
    current_filter_type = ""

    h = headers  # col index by header name

    dim_col = h.get("Dimension")
    attr_col = h.get("Attribute")
    ftype_col = h.get("Filter Type")
    ss_col = h.get("Single Select")
    dv_col = h.get("Default Value")
    mf_col = h.get("Member Filter")  # fallback for default value
    hdm_col = h.get("Hide Dummy Members")
    cal_col = h.get("Calendar Selection")
    sb_col = h.get("Sort By")
    so_col = h.get("Sort Order")

    if not attr_col:
        return filters

    for r in range(start_row, end_row + 1):
        attr_val = _to_str(_cell_val(ws, r, attr_col)) if attr_col else None
        ftype_val = _to_str(_cell_val(ws, r, ftype_col)) if ftype_col else None
        dim_val = _to_str(_cell_val(ws, r, dim_col)) if dim_col else None

        # If Filter Type has a value but Attribute is empty → sub-type label
        if ftype_val and not attr_val:
            current_filter_type = ftype_val
            continue

        # If Attribute has a value → this is a filter data row
        if attr_val:
            if ftype_val:
                current_filter_type = ftype_val

            # Default value: try Default Value column first, fall back to Member Filter
            default_val = _to_str(_cell_val(ws, r, dv_col)) if dv_col else None
            if default_val is None and mf_col:
                default_val = _to_str(_cell_val(ws, r, mf_col))

            filters.append(FilterItem(
                attribute=attr_val,
                dimension=dim_val or "",
                filter_type=current_filter_type,
                single_select=_to_bool(_cell_val(ws, r, ss_col)) if ss_col else None,
                default_value=default_val,
                hide_dummy_members=_to_bool(_cell_val(ws, r, hdm_col)) if hdm_col else None,
                calendar_selection=_to_bool(_cell_val(ws, r, cal_col)) if cal_col else None,
                sort_by=_to_str(_cell_val(ws, r, sb_col)) if sb_col else None,
                sort_order=_to_str(_cell_val(ws, r, so_col)) if so_col else None,
            ))

    return filters


def _parse_dimensions_from_sheet(ws, start_row, end_row, max_col, headers):
    """Parse the dimension section."""
    dims = []
    h = headers

    attr_col = h.get("Attribute")
    dim_col = h.get("Dimension Name")
    axis_col = h.get("Position")
    vis_col = h.get("Visible")
    req_col = h.get("Required")
    sub_col = h.get("Enable Subtotal")
    cond_col = h.get("Show Conditional Format")
    disp_col = h.get("Display Name")
    sort_col = h.get("Sort By")
    type_col = h.get("Type")

    if not attr_col:
        return dims

    for r in range(start_row, end_row + 1):
        attr_val = _to_str(_cell_val(ws, r, attr_col))
        if not attr_val:
            continue

        dims.append(DimensionItem(
            attribute=attr_val,
            dimension=_to_str(_cell_val(ws, r, dim_col)) or "" if dim_col else "",
            axis=_norm_axis(_to_str(_cell_val(ws, r, axis_col))) or "" if axis_col else "",
            visible=_to_bool(_cell_val(ws, r, vis_col)) if vis_col else None,
            required=_to_bool(_cell_val(ws, r, req_col)) if req_col else None,
            show_subtotal=_to_bool(_cell_val(ws, r, sub_col)) if sub_col else None,
            show_conditional_format=_to_bool(_cell_val(ws, r, cond_col)) if cond_col else None,
            display_name=_to_str(_cell_val(ws, r, disp_col)) if disp_col else None,
            sort_by=_to_str(_cell_val(ws, r, sort_col)) if sort_col else None,
            item_type=_to_str(_cell_val(ws, r, type_col)) or "" if type_col else "",
        ))

    return dims


def _parse_measures_from_sheet(ws, start_row, end_row, max_col, headers):
    """Parse the measure section."""
    measures = []
    h = headers

    name_col = h.get("Measure Name")
    vis_col = h.get("Visible")
    edit_col = h.get("Editable")
    rt_col = h.get("Render Type")
    align_col = h.get("Alignment")
    fmt_col = h.get("Format String")
    th_col = h.get("Time Horizon")
    sort_col = h.get("Default Sort")
    tz_col = h.get("Show in Local Timezone")

    if not name_col:
        return measures

    for r in range(start_row, end_row + 1):
        name_val = _to_str(_cell_val(ws, r, name_col))
        if not name_val:
            continue

        measures.append(MeasureItem(
            name=name_val,
            visible=_to_bool(_cell_val(ws, r, vis_col)) if vis_col else None,
            editable=_to_bool(_cell_val(ws, r, edit_col)) if edit_col else None,
            render_type=_to_str(_cell_val(ws, r, rt_col)) if rt_col else None,
            alignment=_to_str(_cell_val(ws, r, align_col)) if align_col else None,
            format_string=_to_str(_cell_val(ws, r, fmt_col)) if fmt_col else None,
            time_horizon=_to_str(_cell_val(ws, r, th_col)) if th_col else None,
            sort=_to_str(_cell_val(ws, r, sort_col)) if sort_col else None,
            show_local_timezone=_to_bool(_cell_val(ws, r, tz_col)) if tz_col else None,
        ))

    return measures


def normalize_wireframe(file_bytes: bytes, sheet_name: Optional[str] = None) -> NormalizedReport:
    """Parse a wireframe XLSX into a NormalizedReport."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    banners = _detect_real_banners(ws, max_row, max_col)

    report = NormalizedReport(source="wireframe")

    for i, (banner_row, section_name) in enumerate(banners):
        section_end = banners[i + 1][0] - 1 if i + 1 < len(banners) else max_row
        header_row = _find_header_row(ws, banner_row + 1, section_end, max_col)
        if not header_row:
            continue

        headers = _read_headers(ws, header_row, max_col)
        data_start = header_row + 1
        name_lower = section_name.lower()

        if "filter" in name_lower:
            report.filters = _parse_filters_from_sheet(
                ws, data_start, section_end, max_col, headers
            )
        elif "dimension" in name_lower:
            report.dimensions = _parse_dimensions_from_sheet(
                ws, data_start, section_end, max_col, headers
            )
        elif "measure" in name_lower:
            report.measures = _parse_measures_from_sheet(
                ws, data_start, section_end, max_col, headers
            )

    return report


# ---------------------------------------------------------------------------
# O9 Report JSON parser
# ---------------------------------------------------------------------------

def normalize_o9_json(data: dict) -> NormalizedReport:
    """Parse an O9 report JSON into a NormalizedReport."""
    report = NormalizedReport(source="o9_report")
    report.report_name = data.get("Name", "")

    # --- Filters: from WidgetModel.ConfigJson.LevelAttributes where IsFilter=true
    wm_attrs = (data.get("WidgetModel") or {}).get("ConfigJson", {}).get("LevelAttributes", [])
    for la in wm_attrs:
        if not la.get("IsFilter", False):
            continue
        members = la.get("SelectedMembers", [])
        default_val = members[0].get("Name", "") if members else None
        report.filters.append(FilterItem(
            attribute=la.get("AttributeName", ""),
            dimension=la.get("Dimension", ""),
            filter_type="",
            single_select=la.get("IsSingleSelect"),
            default_value=default_val,
            hide_dummy_members=la.get("HideDummyMembers"),
            calendar_selection=la.get("CalendarSelection"),
            sort_by=_to_str(la.get("SortBy")),
            sort_order=_to_str(la.get("SortOrder")),
        ))

    # --- Dimensions: from ConfigJson.Presentation.LevelAttributes
    pres_attrs = (data.get("ConfigJson") or {}).get("Presentation", {}).get("LevelAttributes", [])
    for la in pres_attrs:
        report.dimensions.append(DimensionItem(
            attribute=la.get("AttributeName", ""),
            dimension=la.get("Dimension", ""),
            axis=_norm_axis(la.get("Axis")) or "",
            visible=la.get("IsVisible"),
            required=la.get("IsAttributeRequired"),
            show_subtotal=la.get("ShowSubtotal"),
            show_conditional_format=la.get("ShowMemberConditionalFormatting"),
            display_name=_to_str(la.get("DisplayName")),
            sort_by=_to_str(la.get("SortBy")),
        ))

    # --- Measures: from ConfigJson.Presentation.MeasureCollections[].Measures[]
    collections = (data.get("ConfigJson") or {}).get("Presentation", {}).get("MeasureCollections", [])
    for mc in collections:
        for m in mc.get("Measures", []):
            report.measures.append(MeasureItem(
                name=m.get("Name", ""),
                visible=m.get("IsVisible"),
                editable=m.get("IsEditable"),
                render_type=_to_str(m.get("RenderType")),
                alignment=_to_str(m.get("Alignment")),
                format_string=_to_str(m.get("FormatString")),
                time_horizon=_to_str(m.get("TimeHorizon")),
                sort=_to_str(m.get("Sort")),
                show_local_timezone=m.get("ShowInLocalTimeZone"),
            ))

    return report


# ---------------------------------------------------------------------------
# Bidirectional comparison engine
# ---------------------------------------------------------------------------

@dataclass
class DiffRow:
    section: str        # "Filters", "Dimensions", "Measures"
    item_key: str       # The matched item (attribute/name)
    field: str          # Which field differs
    wireframe_value: str
    report_value: str
    status: str         # "match", "mismatch", "only_in_wireframe", "only_in_report"


@dataclass
class ComparisonResult:
    summary: dict
    rows: list
    wireframe_filters: int = 0
    wireframe_dimensions: int = 0
    wireframe_measures: int = 0
    report_filters: int = 0
    report_dimensions: int = 0
    report_measures: int = 0


def _fmt(val: Any) -> str:
    if val is None:
        return "--"
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    return str(val)


def _vals_match(a: Any, b: Any, normalize_fn=None) -> bool:
    """Compare two values, treating None/empty as equivalent."""
    va = a if a is not None else None
    vb = b if b is not None else None

    # Both None/empty
    if va is None and vb is None:
        return True
    if va is None or vb is None:
        # One is None, other is False/empty-string → treat as match
        if va is None and vb in (False, "", 0):
            return True
        if vb is None and va in (False, "", 0):
            return True
        return False

    if normalize_fn:
        va = normalize_fn(va)
        vb = normalize_fn(vb)

    # Bool comparison
    if isinstance(va, bool) or isinstance(vb, bool):
        return _to_bool(va) == _to_bool(vb)

    # String comparison (case-insensitive)
    return str(va).strip().lower() == str(vb).strip().lower()


def _compare_items(section: str, key_label: str,
                   wireframe_items: list, report_items: list,
                   key_fn, field_specs: list) -> list:
    """Compare two lists of items bidirectionally.

    Args:
        section: Section name ("Filters", "Dimensions", "Measures")
        key_label: Human label for the key field
        wireframe_items: Items from wireframe
        report_items: Items from O9 report
        key_fn: Function to extract key from an item
        field_specs: List of (field_name, getter_fn, normalize_fn)

    Returns:
        List of DiffRow
    """
    rows = []

    # Index by key (case-insensitive)
    wf_map = {}
    for item in wireframe_items:
        k = key_fn(item).lower() if key_fn(item) else ""
        wf_map[k] = item

    rp_map = {}
    for item in report_items:
        k = key_fn(item).lower() if key_fn(item) else ""
        rp_map[k] = item

    all_keys = list(dict.fromkeys(
        [key_fn(i) for i in wireframe_items] +
        [key_fn(i) for i in report_items]
    ))

    for key in all_keys:
        k_lower = key.lower() if key else ""
        wf_item = wf_map.get(k_lower)
        rp_item = rp_map.get(k_lower)

        if wf_item and not rp_item:
            rows.append(DiffRow(
                section=section, item_key=key,
                field="(entire item)",
                wireframe_value="Present",
                report_value="--",
                status="only_in_wireframe",
            ))
            continue

        if rp_item and not wf_item:
            rows.append(DiffRow(
                section=section, item_key=key,
                field="(entire item)",
                wireframe_value="--",
                report_value="Present",
                status="only_in_report",
            ))
            continue

        # Both exist — compare field by field
        for field_name, getter, norm_fn in field_specs:
            wf_val = getter(wf_item)
            rp_val = getter(rp_item)
            match = _vals_match(wf_val, rp_val, norm_fn)
            rows.append(DiffRow(
                section=section, item_key=key,
                field=field_name,
                wireframe_value=_fmt(wf_val),
                report_value=_fmt(rp_val),
                status="match" if match else "mismatch",
            ))

    return rows


def compare_reports(wireframe: NormalizedReport, report: NormalizedReport) -> ComparisonResult:
    """Bidirectional field-level comparison of wireframe vs O9 report."""
    all_rows = []

    # --- Filters ---
    filter_specs = [
        ("Dimension", lambda f: f.dimension, None),
        ("Single Select", lambda f: f.single_select, None),
        ("Default Value", lambda f: f.default_value, None),
        ("Hide Dummy Members", lambda f: f.hide_dummy_members, None),
        ("Calendar Selection", lambda f: f.calendar_selection, None),
        ("Sort By", lambda f: f.sort_by, None),
        ("Sort Order", lambda f: f.sort_order, None),
    ]
    all_rows.extend(_compare_items(
        "Filters", "Attribute",
        wireframe.filters, report.filters,
        key_fn=lambda f: f.attribute,
        field_specs=filter_specs,
    ))

    # --- Dimensions ---
    dim_specs = [
        ("Dimension", lambda d: d.dimension, None),
        ("Axis/Position", lambda d: d.axis, _norm_axis),
        ("Visible", lambda d: d.visible, None),
        ("Required", lambda d: d.required, None),
        ("Show Subtotal", lambda d: d.show_subtotal, None),
        ("Show Conditional Format", lambda d: d.show_conditional_format, None),
        ("Display Name", lambda d: d.display_name, None),
        ("Sort By", lambda d: d.sort_by, None),
    ]
    all_rows.extend(_compare_items(
        "Dimensions", "Attribute",
        wireframe.dimensions, report.dimensions,
        key_fn=lambda d: d.attribute,
        field_specs=dim_specs,
    ))

    # --- Measures ---
    measure_specs = [
        ("Visible", lambda m: m.visible, None),
        ("Editable", lambda m: m.editable, None),
        ("Render Type", lambda m: m.render_type, None),
        ("Alignment", lambda m: m.alignment, _norm_align),
        ("Format String", lambda m: m.format_string, None),
        ("Time Horizon", lambda m: m.time_horizon, None),
        ("Sort", lambda m: m.sort, None),
        ("Show Local Timezone", lambda m: m.show_local_timezone, None),
    ]
    all_rows.extend(_compare_items(
        "Measures", "Name",
        wireframe.measures, report.measures,
        key_fn=lambda m: m.name,
        field_specs=measure_specs,
    ))

    # Summary
    total = len(all_rows)
    matches = sum(1 for r in all_rows if r.status == "match")
    mismatches = sum(1 for r in all_rows if r.status == "mismatch")
    only_wf = sum(1 for r in all_rows if r.status == "only_in_wireframe")
    only_rp = sum(1 for r in all_rows if r.status == "only_in_report")

    return ComparisonResult(
        summary={
            "total_fields": total,
            "matches": matches,
            "mismatches": mismatches,
            "only_in_wireframe": only_wf,
            "only_in_report": only_rp,
            "match_percentage": round(matches / total * 100, 1) if total else 100.0,
        },
        rows=[asdict(r) for r in all_rows],
        wireframe_filters=len(wireframe.filters),
        wireframe_dimensions=len(wireframe.dimensions),
        wireframe_measures=len(wireframe.measures),
        report_filters=len(report.filters),
        report_dimensions=len(report.dimensions),
        report_measures=len(report.measures),
    )
