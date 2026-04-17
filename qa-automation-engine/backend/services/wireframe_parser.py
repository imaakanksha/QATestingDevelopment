"""Section-aware XLSX wireframe parser using openpyxl.

Designed for O9 Solutions wireframes which contain multiple sections
(e.g. Filter AOP, Dimensions, Measure) stacked vertically on a single sheet,
separated by colored banner rows with merged cells.

Also handles simple flat-table XLSX files as a fallback.
"""

import io
import math
import re
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Value sanitisation (shared logic)
# ---------------------------------------------------------------------------

def _sanitize_value(val: Any) -> Any:
    """Convert a single cell value to a JSON-safe Python type."""
    if val is None:
        return None

    # Python float: NaN / Inf → None, integral → int
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return None
        if val.is_integer():
            return int(val)
        return val

    # numpy / openpyxl numeric wrappers
    try:
        if hasattr(val, "item"):
            native = val.item()
            if isinstance(native, float):
                if math.isnan(native) or math.isinf(native):
                    return None
                if native.is_integer():
                    return int(native)
            return native
    except (ValueError, OverflowError, AttributeError):
        pass

    # bool (must come before int check since bool is subclass of int)
    if isinstance(val, bool):
        return val

    # int passthrough
    if isinstance(val, int):
        return val

    # datetime / date / time → ISO string
    if hasattr(val, "isoformat") and callable(val.isoformat):
        return val.isoformat()

    # String passthrough
    if isinstance(val, str):
        return val.strip() if val.strip() else None

    return str(val)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_key(text: str) -> str:
    """Normalise a header string to a JSON-friendly key."""
    key = text.strip().lower()
    key = re.sub(r"[^a-z0-9]+", "_", key)
    return key.strip("_")


def _has_fill_color(cell) -> bool:
    """Return True if the cell has a non-white, non-default background fill."""
    fill = cell.fill
    if fill is None or fill.fgColor is None:
        return False
    color = fill.fgColor
    # Theme-based or indexed colours with an actual tint
    if color.type == "theme" and color.theme is not None:
        return True
    if color.type == "rgb" and color.rgb and color.rgb not in ("00000000", "FFFFFFFF"):
        return True
    if color.type == "indexed" and color.indexed is not None and color.indexed not in (0, 64):
        return True
    return False


def _is_merged_wide(ws, row_num: int, min_span: int = 4) -> bool:
    """Return True if any merged-cell range on this row spans ≥ min_span columns."""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row_num <= mr.max_row:
            if (mr.max_col - mr.min_col + 1) >= min_span:
                return True
    return False


def _row_values(ws, row_num: int, max_col: int) -> list:
    """Read all cell values in a row up to max_col."""
    return [ws.cell(row=row_num, column=c).value for c in range(1, max_col + 1)]


def _row_cells(ws, row_num: int, max_col: int) -> list:
    """Read all cell objects in a row up to max_col."""
    return [ws.cell(row=row_num, column=c) for c in range(1, max_col + 1)]


def _count_non_empty(values: list) -> int:
    """Count non-None, non-empty-string values."""
    return sum(1 for v in values if v is not None and (not isinstance(v, str) or v.strip()))


def _is_section_banner(ws, row_num: int, max_col: int) -> Optional[str]:
    """Detect if a row is a coloured section banner.

    Returns the section name string if detected, else None.
    A banner is a row where:
      - The first non-empty cell has a fill colour, OR
      - The row contains a wide merged cell region
      - AND the row has ≤ 2 non-empty cells (just the title)
    """
    values = _row_values(ws, row_num, max_col)
    non_empty = _count_non_empty(values)

    if non_empty == 0:
        return None

    # Find the first non-empty cell
    first_val = None
    first_cell = None
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=c)
        if cell.value is not None and (not isinstance(cell.value, str) or cell.value.strip()):
            first_val = str(cell.value).strip()
            first_cell = cell
            break

    if first_val is None:
        return None

    # Must have only 1–2 non-empty cells (the section title)
    if non_empty > 2:
        return None

    # Check for fill colour OR wide merge
    has_color = _has_fill_color(first_cell) if first_cell else False
    has_merge = _is_merged_wide(ws, row_num, min_span=4)

    if has_color or has_merge:
        return first_val

    return None


def _detect_header_row(ws, start_row: int, end_row: int, max_col: int) -> Optional[int]:
    """Find the header row within a range — the row with the most non-empty text cells."""
    best_row = None
    best_count = 0

    for r in range(start_row, end_row + 1):
        values = _row_values(ws, r, max_col)
        # Count non-empty string cells (headers are text)
        text_count = sum(
            1 for v in values
            if v is not None and isinstance(v, str) and v.strip()
        )
        if text_count > best_count:
            best_count = text_count
            best_row = r

    # Require at least 2 text cells to qualify as a header
    return best_row if best_count >= 2 else None


def _parse_metadata_section(ws, start_row: int, end_row: int, max_col: int) -> dict:
    """Parse the top metadata area as key-value pairs.

    The metadata section typically has labels in some cells and values in adjacent cells,
    e.g.: Row 2: 'Initial ID' | empty | empty | 'Report' | 'Owner/Player' | 'Report Name'
          Row 3: 'Report Description'
    We treat each non-empty cell as a key and the cell immediately to its right as its value
    (if it exists and isn't itself a label in the same pattern).
    For simplicity, we collect all non-empty cells as metadata fields.
    """
    metadata = {}
    for r in range(start_row, end_row + 1):
        row_entries = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and isinstance(val, str) and val.strip():
                row_entries.append((c, val.strip()))

        # Try to pair labels with adjacent values
        used_cols = set()
        for col_idx, label in row_entries:
            if col_idx in used_cols:
                continue
            key = _normalize_key(label)
            # Check if the next column has a value that isn't in our entries list
            next_col = col_idx + 1
            next_val = ws.cell(row=r, column=next_col).value if next_col <= max_col else None
            next_is_label = any(c == next_col for c, _ in row_entries)

            if next_val is not None and not next_is_label:
                metadata[key] = _sanitize_value(next_val)
                used_cols.add(next_col)
            else:
                # Store the label as a key with null value (placeholder)
                if key not in metadata:
                    metadata[key] = None

    return metadata


def _parse_data_section(ws, header_row: int, end_row: int, max_col: int,
                        normalize: bool = True) -> dict:
    """Parse a section with a header row + data rows into column names + records."""
    # Read headers
    raw_headers = _row_values(ws, header_row, max_col)
    headers = []
    col_indices = []
    for i, h in enumerate(raw_headers):
        if h is not None and isinstance(h, str) and h.strip():
            headers.append(h.strip())
            col_indices.append(i + 1)  # 1-based column index

    if not headers:
        return {"columns": [], "data": []}

    # Normalised key names for JSON keys
    keys = [_normalize_key(h) for h in headers] if normalize else headers

    # Read data rows
    records = []
    for r in range(header_row + 1, end_row + 1):
        values = {}
        has_data = False
        for key, col_idx, header in zip(keys, col_indices, headers):
            val = _sanitize_value(ws.cell(row=r, column=col_idx).value)
            values[key] = val
            if val is not None:
                has_data = True

        # Skip completely empty rows
        if has_data:
            records.append(values)

    return {
        "columns": headers,
        "data": records,
    }


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_wireframe_to_json(
    file_bytes: bytes,
    sheet_name: Optional[str] = None,
    normalize_keys: bool = True,
) -> dict:
    """Parse an O9-style wireframe XLSX into structured JSON.

    Detects coloured section banners, extracts metadata from the header area,
    and parses each section (Filter AOP, Dimensions, Measure, etc.) independently.

    Falls back to flat-table parsing if no sections are detected.

    Args:
        file_bytes: Raw bytes of the .xlsx file.
        sheet_name: Sheet to read. None = first sheet.
        normalize_keys: Lowercase + underscore-ify JSON keys.

    Returns:
        Structured dict with 'metadata', 'sections', and 'section_order'.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Select sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # ------------------------------------------------------------------
    # Phase 1: Scan for section banners
    # ------------------------------------------------------------------
    banners = []  # list of (row_number, section_name)
    for r in range(1, max_row + 1):
        section_name = _is_section_banner(ws, r, max_col)
        if section_name:
            banners.append((r, section_name))

    # ------------------------------------------------------------------
    # Phase 2: If no banners found → fallback to flat-table mode
    # ------------------------------------------------------------------
    if not banners:
        header_row = _detect_header_row(ws, 1, min(max_row, 10), max_col)
        if header_row is None:
            header_row = 1

        section = _parse_data_section(ws, header_row, max_row, max_col, normalize_keys)
        return {
            "metadata": {},
            "sections": {
                "data": section,
            },
            "section_order": ["data"],
        }

    # ------------------------------------------------------------------
    # Phase 3: Parse metadata (rows before the first banner)
    # ------------------------------------------------------------------
    first_banner_row = banners[0][0]
    metadata = {}
    if first_banner_row > 1:
        metadata = _parse_metadata_section(ws, 1, first_banner_row - 1, max_col)

    # ------------------------------------------------------------------
    # Phase 4: Parse each section
    # ------------------------------------------------------------------
    sections = {}
    section_order = []

    for i, (banner_row, section_name) in enumerate(banners):
        # Determine the end row for this section
        if i + 1 < len(banners):
            section_end = banners[i + 1][0] - 1
        else:
            section_end = max_row

        # Find header row within the section (between banner and end)
        header_row = _detect_header_row(ws, banner_row + 1, section_end, max_col)
        if header_row is None:
            # No header found — skip this section or store raw
            continue

        # Parse the data section
        section_key = _normalize_key(section_name) if normalize_keys else section_name
        section_data = _parse_data_section(ws, header_row, section_end, max_col, normalize_keys)

        sections[section_key] = section_data
        section_order.append(section_key)

    return {
        "metadata": metadata,
        "sections": sections,
        "section_order": section_order,
    }


def get_wireframe_preview(file_bytes: bytes, sheet_name: Optional[str] = None) -> dict:
    """Quick preview: return detected section names and row counts without full parsing."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    banners = []
    for r in range(1, max_row + 1):
        name = _is_section_banner(ws, r, max_col)
        if name:
            banners.append({"row": r, "name": name})

    return {
        "sheet": ws.title,
        "total_rows": max_row,
        "total_cols": max_col,
        "detected_sections": banners,
        "is_wireframe": len(banners) > 0,
    }
